import openpyxl
import argparse
import os
import sys
import urllib.request
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)),'..'))
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from src.from_response_get_data import single_task_devices_exceptions_summary,single_task_devices_data,single_task_devices_total_run_time,get_devices_list
from Lib.config_parse import parse_config,config_path

DOWNLOAD_HOST = "http://tams.thundersoft.com/go-fastdfs"

def get_tasks_run_time(task_id_list):
    """
    获取所有任务的执行时长，float类型
    :param task_id_list:
    :return:
    """
    tasks_run_time = float(0)
    for task_id in task_id_list:
        tasks_run_time = tasks_run_time + single_task_devices_total_run_time(task_id)
    return tasks_run_time

def get_tasks_devices_exceptions_type_num(task_id_list, witch_version):
    """
    获取所有任务下所有设备的APP_CRASH和ANR两种异常的异常数量和不区分设备的需要用于计算MTBF的APP_CRASH和ANR两种异常的异常数量
    (
        {
            'BH9500G8JG': {
                '+APP_CRASH': 15,
                '+ANR': 0
            },
            'BH95003LJH': {
                '+APP_CRASH': 13,
                '+ANR': 0
            },...
        },
        {
            '+APP_CRASH': 42,
            '+ANR': 4
        }
    )
    :param task_id_list:
    :param witch_version:
    :return:
    """
    tasks_devices_exceptions_type_dict = get_single_device_exceptions_type_num(task_id_list, witch_version)
    tasks_devices_exceptions_type_num_dict = {}
    mtbf_tasks_devices_exceptions_type_num_dict = {}
    for device_id in tasks_devices_exceptions_type_dict.keys():
        tasks_devices_exceptions_type_num_dict[device_id] = {}
        tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] = 0
        tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] = 0
        mtbf_tasks_devices_exceptions_type_num_dict[device_id] = {}
        mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] = 0
        mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] = 0
        exceptions_type_list = []
        for one in tasks_devices_exceptions_type_dict[device_id].keys():
            exceptions_type_list.append(one)
        for i in range(len(exceptions_type_list)):
            if '+APP_CRASH' in exceptions_type_list[i]:
                tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] = \
                    tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] + tasks_devices_exceptions_type_dict[device_id][exceptions_type_list[i]]
            elif '+ANR' in exceptions_type_list[i]:
                tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] = \
                    tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] + tasks_devices_exceptions_type_dict[device_id][exceptions_type_list[i]]
            if witch_version == 'version_213':
                if '+APP_CRASH' in exceptions_type_list[i] and 'sony' in exceptions_type_list[i]:
                    mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] = \
                        mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH'] + \
                        tasks_devices_exceptions_type_dict[device_id][exceptions_type_list[i]]
                elif '+ANR' in exceptions_type_list[i] and 'sony' in exceptions_type_list[i]:
                    mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] = \
                        mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+ANR'] + tasks_devices_exceptions_type_dict[device_id][exceptions_type_list[i]]
            elif witch_version == 'version_225':
                mtbf_tasks_devices_exceptions_type_num_dict = tasks_devices_exceptions_type_num_dict
    # two_type_exceptions_num_dict = {}
    # two_type_exceptions_num_dict['+APP_CRASH'] = 0
    # two_type_exceptions_num_dict['+ANR'] = 0
    # for device_id in tasks_devices_exceptions_type_num_dict.keys():
    #     if tasks_devices_exceptions_type_num_dict[device_id].get('+APP_CRASH'):
    #         two_type_exceptions_num_dict['+APP_CRASH'] = \
    #             two_type_exceptions_num_dict['+APP_CRASH'] + tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH']
    #     elif tasks_devices_exceptions_type_num_dict[device_id].get('+ANR'):
    #         two_type_exceptions_num_dict['+ANR'] = \
    #             two_type_exceptions_num_dict['+ANR'] + tasks_devices_exceptions_type_num_dict[device_id]['+ANR']
    mtbf_exceptions_type_num_dict = {}
    mtbf_exceptions_type_num_dict['+APP_CRASH'] = 0
    mtbf_exceptions_type_num_dict['+ANR'] = 0
    for device_id in mtbf_tasks_devices_exceptions_type_num_dict.keys():
        if mtbf_tasks_devices_exceptions_type_num_dict[device_id].get('+APP_CRASH'):
            mtbf_exceptions_type_num_dict['+APP_CRASH'] = \
                mtbf_exceptions_type_num_dict['+APP_CRASH'] + mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+APP_CRASH']
        elif mtbf_tasks_devices_exceptions_type_num_dict[device_id].get('+ANR'):
            mtbf_exceptions_type_num_dict['+ANR'] = \
                mtbf_exceptions_type_num_dict['+ANR'] + mtbf_tasks_devices_exceptions_type_num_dict[device_id]['+ANR']
    # return tasks_devices_exceptions_type_num_dict,two_type_exceptions_num_dict,mtbf_exceptions_type_num_dict
    return tasks_devices_exceptions_type_num_dict,mtbf_exceptions_type_num_dict

def get_all_tasks_devices_test_result(task_id_list):
    """
    获取所有任务下所有设备的test_time，pass_case，total_case，testcase_count，pass_rate
    {
        'BH9500G8JG': {
            'test_time': 4.85,      # 执行时长：            4.85h
            'pass_case': 143,       # 测试通过的case数：     143
            'total_case': 161,      # 测试的case总数：      161
            'testcase_count': 18,   # 测试的case个数：      18
            'pass_rate': '88.82%'    # 测试的case通过率：     89%
        },
        'BH95003LJH': {
            'test_time': 5.03,
            'pass_case': 131,
            'total_case': 161,
            'testcase_count': 18,
            'pass_rate': '81.37%'
        },...
    }
    :param task_id_list:
    :return:
    """
    all_tasks_devices_test_result = {}
    for task_id in task_id_list:
        single_task_devices_data_dict = single_task_devices_data(task_id)
        for device_id in single_task_devices_data_dict.keys():
            if all_tasks_devices_test_result.get(device_id):
                all_tasks_devices_test_result[device_id]['test_time'] = \
                    all_tasks_devices_test_result[device_id]['test_time'] + single_task_devices_data_dict[device_id]['test_time']
                all_tasks_devices_test_result[device_id]['pass_case'] = \
                    all_tasks_devices_test_result[device_id]['pass_case'] + single_task_devices_data_dict[device_id]['pass_case']
                all_tasks_devices_test_result[device_id]['total_case'] = \
                    all_tasks_devices_test_result[device_id]['total_case'] + single_task_devices_data_dict[device_id]['total_case']
                all_tasks_devices_test_result[device_id]['testcase_count'] = \
                    max(all_tasks_devices_test_result[device_id]['testcase_count'],single_task_devices_data_dict[device_id]['testcase_count'])
                all_tasks_devices_test_result[device_id]['pass_rate'] = \
                    str(round(float((all_tasks_devices_test_result[device_id]['pass_case'] + single_task_devices_data_dict[device_id]['pass_case'])/(all_tasks_devices_test_result[device_id]['total_case'] + single_task_devices_data_dict[device_id]['total_case']))*100,2)) + '%'
            else:
                all_tasks_devices_test_result[device_id] = single_task_devices_data_dict[device_id]
    return all_tasks_devices_test_result

def get_different_type_exception_num(task_id_list,witch_version):
    """
    不同类型的缺陷的数目,返回的是一个元组,键值对是exception类型:exception数量
    (
        {
            'com.sonymobile.audioutil+APP_CRASH': 42,
            'com.sonymobile.launcher+ANR': 4
        },
        46,
        2
    )
    :param task_id_list:
    :param witch_version:
    :return:
    """
    exception_type_num_dict = {}
    exceptions_num = 0
    mtbf_exceptions_num = 0
    tasks_devices_exceptions_type_list, tasks_devices_exceptions_summary_dict = get_all_tasks_devices_exceptions_type(task_id_list, witch_version)
    for one in tasks_devices_exceptions_type_list:
        exception_type_num_dict[one] = 0
        for device_id in tasks_devices_exceptions_summary_dict.keys():
            for package_name in tasks_devices_exceptions_summary_dict[device_id]['package_name']:
                if one in package_name:
                    exception_type_num_dict[one] = exception_type_num_dict[one] + 1
    for exception_type in exception_type_num_dict.keys():
        exceptions_num = exceptions_num + exception_type_num_dict[exception_type]
    # if exceptions_num == 0:
    #     print(f'当前任务{task_id_list}下exception数量为0')
    #     exceptions_num = 1
    if witch_version == 'version_213':
        for exception in exception_type_num_dict.keys():
            if 'sony' in exception:
                mtbf_exceptions_num = mtbf_exceptions_num + exception_type_num_dict[exception]
    #     if mtbf_exceptions_num == 0:
    #         mtbf_exceptions_num = 1
    # else:
    #     print(f'当前任务{task_id_list}下包含"sony"字样的exception数量为0')
    #     mtbf_exceptions_num = 1
    if witch_version == 'version_225':
        mtbf_exceptions_num = exceptions_num
    return exception_type_num_dict,exceptions_num,mtbf_exceptions_num

def get_all_tasks_devices_list(task_id_list):
    """
    获取所有任务所有设备的不重复集合（列表）
    ['BH9500G8JG', 'BH95003LJH', 'BH95002WJH', 'BH9500S7JG', 'BH9500YMJG', 'HQ60BK3089', 'HQ60CN3067', 'HQ60BK3096', 'HQ60BK3092', 'HQ60BK3902']
    :param task_id_list:
    :return:
    """
    device_id_list_summary = []
    for task_id in task_id_list:
        for device_id in get_devices_list(task_id):
            if device_id not in device_id_list_summary:
                device_id_list_summary.append(device_id)
    return device_id_list_summary

def get_all_tasks_devices_exceptions_summary(task_id_list,witch_version):
    """
    获取所有任务所有设备下的exceptions集合（字典），以device_id为键
    {
    	'BH9500S7JG': {
            'log_path': ['/group1/default/20220113/14/57/3/logcat_20220113145655'],
            'package_name': ['com.sonymobile.launcher+ANR'],
            'filename': ['logcat_20220113145655']
        },
        'BH9500YMJG': {
            'log_path': ['/group1/default/20220113/14/58/3/logcat_20220113145751', '/group1/default/20220113/15/36/3/logcat_20220113153535', '/group1/default/20220113/15/57/3/logcat_20220113155642'],
            'package_name': ['com.sonymobile.launcher+ANR', 'com.sonymobile.launcher+ANR', 'com.sonymobile.launcher+ANR'],
            'filename': ['logcat_20220113145751', 'logcat_20220113153535', 'logcat_20220113155642']
        },...
    }
    :param task_id_list:
    :param witch_version:
    :return:
    """
    tasks_devices_exceptions_summary_dict = {}
    for task_id in task_id_list:
        devices_list = get_devices_list(task_id)
        device_exceptions = single_task_devices_exceptions_summary(task_id, witch_version)
        for device in devices_list:
            if device not in tasks_devices_exceptions_summary_dict.keys():
                tasks_devices_exceptions_summary_dict[device] = device_exceptions[device]
            else:
                for i in range(len(device_exceptions[device]['package_name'])):
                    tasks_devices_exceptions_summary_dict[device]['log_path'].append(device_exceptions[device]['log_path'][i])
                    tasks_devices_exceptions_summary_dict[device]['package_name'].append(device_exceptions[device]['package_name'][i])
                    tasks_devices_exceptions_summary_dict[device]['filename'].append(device_exceptions[device]['filename'][i])
    return tasks_devices_exceptions_summary_dict

def get_all_tasks_devices_exceptions_type(task_id_list,witch_version):
    """
    获取所有任务所有设备下的exceptions不重复类型的集合（列表）
    (
        ['com.sonymobile.launcher+ANR', ' com.sonymobile.audioutil+APP_CRASH'],

        {'BH9500S7JG': {
            'log_path': ['/group1/default/20220113/14/57/3/logcat_20220113145655'],
            'package_name': ['com.sonymobile.launcher+ANR'],
            'filename': ['logcat_20220113145655']
        },
        'BH9500YMJG': {
            'log_path': ['/group1/default/20220113/14/58/3/logcat_20220113145751', '/group1/default/20220113/15/36/3/logcat_20220113153535', '/group1/default/20220113/15/57/3/logcat_20220113155642'],
            'package_name': ['com.sonymobile.launcher+ANR', 'com.sonymobile.launcher+ANR', 'com.sonymobile.launcher+ANR'],
            'filename': ['logcat_20220113145751', 'logcat_20220113153535', 'logcat_20220113155642']
        },...
    )
    :param task_id_list:
    :param witch_version:
    :return:
    """
    tasks_devices_exceptions_type_list = []
    tasks_devices_exceptions_summary_dict = get_all_tasks_devices_exceptions_summary(task_id_list, witch_version)
    for key in tasks_devices_exceptions_summary_dict:
        for exception_type in tasks_devices_exceptions_summary_dict[key]['package_name']:
            tasks_devices_exceptions_type_list.append(exception_type)
    tasks_devices_exceptions_type_list = list(set(tasks_devices_exceptions_type_list))
    return tasks_devices_exceptions_type_list,tasks_devices_exceptions_summary_dict

def get_single_device_exceptions_type_num(task_id_list,witch_version):
    """
    获取所有任务所有设备下不同类型的exception的数目，汇总成一个集合（字典）
    {
        'BH9500G8JG': {
            'com.sonymobile.audioutil+APP_CRASH': 15,
            'com.sonymobile.launcher+ANR': 0
        },
        'BH95003LJH': {
            'com.sonymobile.audioutil+APP_CRASH': 13,
            'com.sonymobile.launcher+ANR': 0
        },...
	}
    :param task_id_list:
    :param witch_version:
    :return:
    """
    tasks_devices_exceptions_type_list,tasks_devices_exceptions_summary_dict = get_all_tasks_devices_exceptions_type(task_id_list, witch_version)
    device_id_list_summary = get_all_tasks_devices_list(task_id_list)
    tasks_devices_exceptions_type_dict = {}
    for device_id in device_id_list_summary:
        single_device_exceptions_type_dict = {device_id: {}}
        for exception in tasks_devices_exceptions_type_list:
            if not tasks_devices_exceptions_type_dict.get(device_id):
                single_device_exceptions_type_dict[device_id][exception] = tasks_devices_exceptions_summary_dict[device_id]['package_name'].count(exception)
            else:
                single_device_exceptions_type_dict[device_id][exception] = single_device_exceptions_type_dict[device_id][exception] + tasks_devices_exceptions_summary_dict[device_id]['package_name'].count(exception)
        tasks_devices_exceptions_type_dict = { **tasks_devices_exceptions_type_dict , **single_device_exceptions_type_dict}
    return tasks_devices_exceptions_type_dict

def download_logfile(task_id,device_id,exception_type,exception_name,file_name,log_path):
    """
    下载logcat文件到本地
    :param task_id:taskID
    :param device_id: deviceID
    :param exception_type: ANR/APP-CRASH
    :param exception_name: com.sonymobile.launcher(package_name)
    :param file_name:logcat_20220113145751
    :param log_path:/group1/default/20220113/14/58/3/logcat_20220113145751
    :return:
    """
    # url = 'http://tams.thundersoft.com/api/storage/download//group1/default/20220113/15/06/3/logcat_20220113150607'
    url = DOWNLOAD_HOST + f'api/storage/download/{log_path}'
    if_pyinstall = parse_config('tams', config_path)['if_pyinstall']
    if if_pyinstall == 'not_pyinstall':
        local_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'product',f'{task_id}',f'{device_id}', f'{exception_type}', f'{exception_name}')
    else:
        # Pyinstaller 之后需要根据执行路径去获取文件下载路径
        local_path = os.path.join(os.path.dirname(sys.argv[0]), 'product',f'{task_id}', f'{device_id}',f'{exception_type}', f'{exception_name}')
    # Create local device_id-exception_type-exception_name folder
    if not os.path.exists(local_path):
        os.makedirs(local_path)
    urllib.request.urlretrieve(url,os.path.join(local_path,f'{file_name}'))
    return True

if __name__ == '__main__':
    # Param init
    parse = argparse.ArgumentParser()
    parse.add_argument("-t", "--task_id", help="The id of task in tams, multiple ids are separated by commas ','",
                       type=str, default="")
    parse.add_argument("-m", "--merge", help="Whether to merge multiple task results, "
                                             "only tasks of the same type are supported ", type=bool, default=False)
    args = parse.parse_args()
    task_id, merge = args.task_id, args.merge

    if not task_id:
        print("[ERROR] Missing key parameter -t or --task_id")
        sys.exit(1)

    # # Multitasking
    # task_id_list = task_id.split(",")

    if merge:
        task_ids_all = []
        task_ids = task_id.split(",")
        task_ids_all.append(task_ids)
    else:
        task_ids_all = []
        task_ids = task_id.split(",")
        for one in task_ids:
            one_list = []
            one_list.append(one)
            task_ids_all.append(one_list)

    for task_id_list in task_ids_all:
        # task_id_list = ['101318', '101744']
        # task_id_list = ['101318', '101318']
        witch_version = parse_config('tams', config_path)['version']
        # 数据导入到excel表格中
        template_excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'product', 'mtbf_report_template.xlsx')    # 获取模板excel文件绝对路径
        # # Pyinstaller 之后需要根据执行路径去获取excel文件路径
        # template_excel_path = os.path.join(os.path.dirname(sys.argv[0]), 'product','mtbf_report_template.xlsx')  # 获取模板excel文件绝对路径
        device_id_list_summary = get_all_tasks_devices_list(task_id_list)  # 获取所有任务下的不同device的列表
        nums_of_devices = len(device_id_list_summary)  # 获取所有任务下的不同device的数量
        work_book = openpyxl.load_workbook(template_excel_path)     # 打开excel文件
        # Sheet页操作：Summary-Sheet页写入数据
        work_sheet_summary = work_book['Summary']   # 打开Summary-sheet页
        # 第4行写入Variant
        work_sheet_summary.cell(row=4, column=3, value="")  # 清空Variant
        work_sheet_summary.cell(row=4, column=3, value="SP1")  # Variant写入固定值SP1
        # 第7行写入RunTime (Hrs)
        tasks_run_time = get_tasks_run_time(task_id_list)
        work_sheet_summary.cell(row=7, column=8, value=tasks_run_time)  # RunTime (Hrs)写入
        # 第10行写入数据
        work_sheet_summary.cell(row=10,column=1,value=nums_of_devices)    # Device Quantity
        work_sheet_summary.cell(row=10, column=2, value=tasks_run_time)  # Totally Run time(Hrs)
        work_sheet_summary.cell(row=10, column=3, value=0)  # Native-Crash Count(Tomstone)
        exception_type_num_dict,exceptions_num,mtbf_exceptions_num = get_different_type_exception_num(task_id_list, witch_version)
        tasks_devices_exceptions_type_num_dict,mtbf_exceptions_type_num_dict = get_tasks_devices_exceptions_type_num(task_id_list, witch_version)
        if mtbf_exceptions_type_num_dict.get('+APP_CRASH'):   # 如果有APP_CRASH这类exception，就按照统计值写入
            work_sheet_summary.cell(row=10, column=4, value=mtbf_exceptions_type_num_dict['+APP_CRASH'])   # APK-Crash Count
        else:   # 如果没有APP_CRASH这类exception，就写入0
            work_sheet_summary.cell(row=10, column=4, value=0)  # APK-Crash Count
        if mtbf_exceptions_type_num_dict.get('+ANR'):   # 如果有ANR这类exception，就按照统计值写入
            work_sheet_summary.cell(row=10, column=5, value=mtbf_exceptions_type_num_dict['+ANR'])   # APK-ANR Count
        else:   # 如果没有ANR这类exception，就写入0
            work_sheet_summary.cell(row=10, column=5, value=0)  # APK-ANR Count
        work_sheet_summary.cell(row=10, column=6, value=mtbf_exceptions_num)  # Total
        if mtbf_exceptions_num != 0:
            work_sheet_summary.cell(row=10, column=7, value=round(float(tasks_run_time/mtbf_exceptions_num),2))  # MTBF
        elif mtbf_exceptions_num == 0:
            work_sheet_summary.cell(row=10, column=7, value=round(float(tasks_run_time), 2))  # MTBF
        # 第13行开始写入每台设备的数据
        tasks_devices_exceptions_type_dict = get_single_device_exceptions_type_num(task_id_list,witch_version)
        tasks_devices_exceptions_type_list,tasks_devices_exceptions_summary_dict = get_all_tasks_devices_exceptions_type(task_id_list, witch_version)
        all_tasks_devices_test_result = get_all_tasks_devices_test_result(task_id_list)
        if nums_of_devices > 1:
            work_sheet_summary.insert_rows(13,nums_of_devices - 1)  # 插入设备数的行
        for i in range(nums_of_devices):
            work_sheet_summary.cell(row=13 + i, column=1, value=device_id_list_summary[i])  # 每行插入Device ID
            work_sheet_summary.cell(row=13 + i, column=2, value=all_tasks_devices_test_result[device_id_list_summary[i]]['test_time'])  # 每行插入Totally Run time(Hrs)
            work_sheet_summary.cell(row=13 + i, column=3, value=all_tasks_devices_test_result[device_id_list_summary[i]]['total_case'])  # 每行插入Total Case
            work_sheet_summary.cell(row=13 + i, column=4, value=all_tasks_devices_test_result[device_id_list_summary[i]]['pass_rate'])  # 每行插入Pass Rate
            work_sheet_summary.cell(row=13 + i, column=5, value=0)  # 每行插入Native-Crash(Tomstone)
            work_sheet_summary.cell(row=13 + i, column=6, value=tasks_devices_exceptions_type_num_dict[device_id_list_summary[i]]['+APP_CRASH'])  # 每行插入APP-Crash
            work_sheet_summary.cell(row=13 + i, column=7, value=tasks_devices_exceptions_type_num_dict[device_id_list_summary[i]]['+ANR'])  # 每行插入APP-ANR
            work_sheet_summary.cell(row=13 + i, column=8, value=tasks_devices_exceptions_type_num_dict[device_id_list_summary[i]]['+APP_CRASH'] + tasks_devices_exceptions_type_num_dict[device_id_list_summary[i]]['+ANR'])  # 每行插入Total

        # Sheet页操作：DetailedInfo_Each Device-Sheet页写入数据
        work_sheet_detail = work_book['DetailedInfo_Each Device']  # 打开Summary-sheet页
        # 每列插入device_id
        work_sheet_detail.insert_cols(5,nums_of_devices - 1)    # 插入设备数的列
        for i in range(nums_of_devices):
            work_sheet_detail.cell(row=1, column=4 + i, value=device_id_list_summary[i])  # 每列插入device_id
        # 每行插入exception名称和类型
        for i in range(len(tasks_devices_exceptions_type_list)):
            work_sheet_detail.cell(row=2 + i, column=1, value=tasks_devices_exceptions_type_list[i].split('+')[0])  # 每行插入exception名称(Package)
            work_sheet_detail.cell(row=2 + i, column=2,value=tasks_devices_exceptions_type_list[i].split('+')[-1])  # 每行插入exception类型(Bug Category)
            work_sheet_detail.cell(row=2 + i, column=nums_of_devices + 4, value=exception_type_num_dict[tasks_devices_exceptions_type_list[i]])  # 每行插入exception类型的Total
        # 每个单元格插入某台设备某个exception的数目
        for i in range(len(tasks_devices_exceptions_type_list)):
            for j in range(nums_of_devices):
                work_sheet_detail.cell(row=2 + i, column=4 + j, value=tasks_devices_exceptions_type_dict[device_id_list_summary[j]][tasks_devices_exceptions_type_list[i]])  # 每个单元格插入某台设备某个exception的数目
        # 保存写入的数据
        if_pyinstall = parse_config('tams', config_path)['if_pyinstall']
        if merge:
            if if_pyinstall == 'not_pyinstall':
                report_excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'product','mtbf_report_template_merge.xlsx')  # 获取最终保存的文件的excel文件绝对路径
            else:
                # Pyinstaller 之后需要根据执行路径去获取excel文件路径
                report_excel_path = os.path.join(os.path.dirname(sys.argv[0]), 'product','mtbf_report_template_merge.xlsx')   # 获取模板excel文件绝对路径
        else:
            if if_pyinstall == 'not_pyinstall':
                report_excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'product',f'mtbf_report_template_{task_id_list}.xlsx')  # 获取最终保存的文件的excel文件绝对路径
            else:
                # Pyinstaller 之后需要根据执行路径去获取excel文件路径
                report_excel_path = os.path.join(os.path.dirname(sys.argv[0]), 'product',f'mtbf_report_template_{task_id_list}.xlsx')  # 获取模板excel文件绝对路径
        work_book.save(report_excel_path)

        # 下载logcat文件到本地
        download_limit = parse_config('tams', config_path)['download_limit']
        download_log_switch = parse_config('tams', config_path)['download_log_switch']
        # download_logfile('BH95002WJH', 'APP_CRASH', 'com.sonymobile.audioutil', 'logcat_20220113150607','/group1/default/20220113/15/06/3/logcat_20220113150607')
        if download_log_switch == "ON":
            for device_id in device_id_list_summary:
                for i in range(len(tasks_devices_exceptions_type_list)):
                    exception_type = tasks_devices_exceptions_type_list[i].split('+')[1]
                    exception_name = tasks_devices_exceptions_type_list[i].split('+')[0]
                    num = tasks_devices_exceptions_type_dict[device_id][tasks_devices_exceptions_type_list[i]]
                    if num >= int(download_limit):
                        index_list = []
                        for tag in range(int(download_limit)):
                            index = tasks_devices_exceptions_summary_dict[device_id]['package_name'].index(tasks_devices_exceptions_type_list[i])
                            index_list.append(index)
                            for j in range(int(download_limit) - 1):
                                index = tasks_devices_exceptions_summary_dict[device_id]['package_name'].index(tasks_devices_exceptions_type_list[i], index + 1)
                                index_list.append(index)
                            for index in index_list:
                                file_name = tasks_devices_exceptions_summary_dict[device_id]['filename'][index]
                                log_path = tasks_devices_exceptions_summary_dict[device_id]['log_path'][index]
                                if merge:
                                    download_logfile('merge',device_id, exception_type, exception_name, file_name, log_path)
                                else:
                                    download_logfile(f'{task_id_list}',device_id, exception_type, exception_name, file_name, log_path)
                    elif num >=1:
                        index_list = []
                        for tag in range(num):
                            index = tasks_devices_exceptions_summary_dict[device_id]['package_name'].index(tasks_devices_exceptions_type_list[i])
                            index_list.append(index)
                            for j in range(num - 1):
                                index = tasks_devices_exceptions_summary_dict[device_id]['package_name'].index(tasks_devices_exceptions_type_list[i], index + 1)
                                index_list.append(index)
                            for index in index_list:
                                file_name = tasks_devices_exceptions_summary_dict[device_id]['filename'][index]
                                log_path = tasks_devices_exceptions_summary_dict[device_id]['log_path'][index]
                                if merge:
                                    download_logfile('merge', device_id, exception_type, exception_name, file_name,log_path)
                                else:
                                    download_logfile(f'{task_id_list}', device_id, exception_type, exception_name,file_name, log_path)


            # 此处可作为下次优化的逻辑（下载logcat文件）
            # device_info = tasks_devices_exceptions_summary_dict.get(device_id)
            # package_list, path_list = device_info["package_name"], device_info["log_path"]
            # print(package_list, path_list)
            # for i in range(len(path_list)):
            #     name, t = package_list[i].split("+")
            #     filename = os.path.basename(path_list[i])
            #     print(device_id, name, t, filename)
            #     print("===========================")
